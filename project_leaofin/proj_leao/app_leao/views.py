import pandas as pd
import io
import json
from datetime import date, datetime, timedelta
from dateutil.relativedelta import relativedelta
from decimal import Decimal, InvalidOperation
import openpyxl
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.core.paginator import Paginator
from django.db import transaction  # <-- ADICIONADO (para transaction.atomic)
from django.db.models import (
    Count, 
    Q, 
    Sum, 
    Value, 
    FloatField, 
    DecimalField  # <-- ADICIONADO (caso precise tratar Decimal no banco)
)
from django.db.models.functions import Coalesce
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from app_leao.models import (
    BancoSaldo,
    Categoria,
    ConciliacaoBancaria,
    ContaPagar,
    Fornecedor,
    TransacaoExtrato,
    FechamentoCaixa,
    Deposito
)

try:
    from ofxtools.Parser import OFXTree
except ImportError:
    OFXTree = None
from django.views.decorators.http import require_POST

STATUS_VALIDOS = [choice[0] for choice in ContaPagar.STATUS_CHOICES]

def extrair_texto(celula):
    """Garante que qualquer valor de célula (int, float, None) vire string limpa."""
    if celula is None:
        return ""
    return str(celula).strip()


def limpar_cnpj(val):
    """Remove caracteres não numéricos do CNPJ."""
    return re.sub(r'\D', '', extrair_texto(val))


def formatar_cnpj(c):
    """Insere a pontuação padrão no CNPJ limpo (ex: 00.000.000/0001-00)."""
    if len(c) == 14:
        return f"{c[:2]}.{c[2:5]}.{c[5:8]}/{c[8:12]}-{c[12:]}"
    return c


def parse_valor(val):
    """Converte valores em Decimal do Django tratando formatação BRL."""
    if isinstance(val, (int, float)):
        return Decimal(str(val))
    texto = extrair_texto(val).replace('R$', '').replace('.', '').replace(',', '.').strip()
    return Decimal(texto) if texto else Decimal('0.00')


def parse_data(val):
    """Trata datas do Excel (datetime, date, serial de dias do Excel ou vários formatos de texto)."""
    if not val:
        return None

    # 1. Se já for objeto datetime ou date do Python
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val

    # 2. Se for número de série do Excel (ex: 45123)
    if isinstance(val, (int, float)):
        try:
            return openpyxl.utils.datetime.from_excel(val).date()
        except Exception:
            pass

    # 3. Se for string, limpa e testa os formatos mais comuns
    texto = extrair_texto(val).split(' ')[0] # Pega só a data se tiver hora junta
    
    formatos = [
        '%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d.%m.%Y',
        '%d/%m/%y', '%Y/%m/%d', '%m/%d/%Y'
    ]

    for fmt in formatos:
        try:
            return datetime.strptime(texto, fmt).date()
        except ValueError:
            pass

    return None

@require_POST
def importar_xlsx(request):
    excel_file = request.FILES.get('arquivo_xlsx') or request.FILES.get('arquivo_excel')
    
    if not excel_file:
        return JsonResponse({'sucesso': False, 'erro': 'Nenhum arquivo enviado.'}, status=400)

    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        sheet = wb.active
        
        # Cabeçalho padronizado em minúsculas
        header = [extrair_texto(cell.value).lower() for cell in sheet[1]]
        
        def achar_coluna(nomes):
            for nome in nomes:
                if nome in header:
                    return header.index(nome)
            return None

        # Mapeamento dinâmico das colunas
        idx_cnpj = achar_coluna(['cnpj', 'cnpj fornecedor', 'cpf/cnpj', 'fornecedor_cnpj'])
        idx_nf = achar_coluna(['nota fiscal', 'nf', 'numero nf', 'num_nota', 'nota_fiscal'])
        idx_linha = achar_coluna(['linha digitavel', 'boleto', 'codigo de barras', 'linha_digitavel'])
        idx_valor = achar_coluna(['valor', 'valor (r$)', 'valor total', 'valor_total'])
        idx_venc = achar_coluna(['vencimento', 'data vencimento', 'dt vencimento', 'data_vencimento', 'venc'])
        idx_categoria = achar_coluna(['categoria', 'cat', 'categoria_nome'])
        idx_banco = achar_coluna(['banco', 'conta bancaria', 'banco_nome', 'bancosaldo'])
        idx_parcela = achar_coluna(['parcela', 'nº parcela', 'numero parcela', 'parcela_numero'])

        # Validação de presença das colunas obrigatórias no cabeçalho
        colunas_faltantes = []
        if idx_cnpj is None: colunas_faltantes.append('CNPJ')
        if idx_valor is None: colunas_faltantes.append('Valor')
        if idx_venc is None: colunas_faltantes.append('Vencimento')
        if idx_categoria is None: colunas_faltantes.append('Categoria')
        if idx_banco is None: colunas_faltantes.append('Banco')
        if idx_parcela is None: colunas_faltantes.append('Parcela')
        if colunas_faltantes:
            return JsonResponse({
                'sucesso': False, 
                'erro': f'Planilha inválida. As colunas a seguir são obrigatórias: {", ".join(colunas_faltantes)}.'
            }, status=400)

        contas_novas = []
        contas_atualizadas = 0
        erros = []

        with transaction.atomic():
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row): 
                    continue

                # --- 1. VALIDAR FORNECEDOR ---
                cnpj_raw = row[idx_cnpj] if idx_cnpj < len(row) else None
                cnpj_limpo = limpar_cnpj(cnpj_raw)

                if not cnpj_limpo:
                    erros.append(f"Linha {row_idx}: CNPJ ausente ou inválido.")
                    continue

                cnpj_formatado = formatar_cnpj(cnpj_limpo)
                fornecedor = Fornecedor.objects.filter(cnpj__in=[cnpj_limpo, cnpj_formatado]).first()
                if not fornecedor:
                    fornecedor = Fornecedor.objects.filter(cnpj__icontains=cnpj_limpo).first()

                if not fornecedor:
                    erros.append(f"Linha {row_idx}: Fornecedor com CNPJ {cnpj_limpo} não encontrado.")
                    continue

                # --- 2. VALIDAR CATEGORIA ---
                nome_categoria = extrair_texto(row[idx_categoria] if idx_categoria < len(row) else "")
                if not nome_categoria:
                    erros.append(f"Linha {row_idx}: Categoria não informada.")
                    continue

                categoria = Categoria.objects.filter(nome__iexact=nome_categoria).first()
                if not categoria:
                    erros.append(f"Linha {row_idx}: Categoria '{nome_categoria}' não encontrada no sistema.")
                    continue

                # --- 3. VALIDAR BANCOSALDO ---
                nome_banco = extrair_texto(row[idx_banco] if idx_banco < len(row) else "")
                if not nome_banco:
                    erros.append(f"Linha {row_idx}: Banco não informado.")
                    continue

                banco_saldo = None
                if hasattr(BancoSaldo, 'nome'):
                    banco_saldo = BancoSaldo.objects.filter(nome__icontains=nome_banco).first()
                if not banco_saldo and hasattr(BancoSaldo, 'descricao'):
                    banco_saldo = BancoSaldo.objects.filter(descricao__icontains=nome_banco).first()
                if not banco_saldo and hasattr(BancoSaldo, 'banco'):
                    banco_saldo = BancoSaldo.objects.filter(banco__icontains=nome_banco).first()
                if not banco_saldo and nome_banco.isdigit():
                    banco_saldo = BancoSaldo.objects.filter(pk=nome_banco).first()

                if not banco_saldo:
                    erros.append(f"Linha {row_idx}: Banco/Conta '{nome_banco}' não encontrado no sistema.")
                    continue

                # --- 4. VALIDAR VALOR E VENCIMENTO ---
                valor = parse_valor(row[idx_valor] if idx_valor < len(row) else 0)
                vencimento = parse_data(row[idx_venc] if idx_venc < len(row) else None)
                parcela = extrair_texto(row[idx_parcela] if idx_parcela < len(row) else "01/01")

                if not vencimento:
                    erros.append(f"Linha {row_idx}: Data de vencimento inválida.")
                    continue

                # --- 5. DADOS OPCIONAIS ---
                nf = extrair_texto(row[idx_nf]) if idx_nf is not None and idx_nf < len(row) else ""
                linha_dig = extrair_texto(row[idx_linha]) if idx_linha is not None and idx_linha < len(row) else ""
                parcela = extrair_texto(row[idx_parcela]) if idx_parcela is not None and idx_parcela < len(row) else "01/01"

                # --- 6. CHECK DE DUPLICIDADE E ATUALIZAÇÃO PARCIAL ---
                # Ajuste os nomes dos atributos 'vencimento', 'nota_fiscal' ou 'banco' 
                # se no seu modelo ContaPagar estiverem nomeados diferente!
                conta_existente = ContaPagar.objects.filter(
                    fornecedor=fornecedor,
                    vencimento=vencimento,
                    valor=valor
                ).first()

                if conta_existente:
                    atualizou = False

                    if not conta_existente.nota_fiscal and nf:
                        conta_existente.nota_fiscal = nf
                        atualizou = True

                    if not conta_existente.linha_digitavel and linha_dig:
                        conta_existente.linha_digitavel = linha_dig
                        atualizou = True

                    if atualizou:
                        conta_existente.save()
                        contas_atualizadas += 1
                    continue

                # --- 7. CRIAR NOVA CONTA ---
                contas_novas.append(ContaPagar(
                    fornecedor=fornecedor,
                    categoria=categoria,
                    banco=banco_saldo,          # Atribui o objeto BancoSaldo
                    nota_fiscal=nf,
                    linha_digitavel=linha_dig,
                    valor=valor,
                    vencimento=vencimento,
                    status='Pendente',
                    parcela=parcela
                ))

            if contas_novas:
                ContaPagar.objects.bulk_create(contas_novas)

        return JsonResponse({
            'sucesso': True,
            'importados': len(contas_novas),
            'atualizados': contas_atualizadas,
            'erros': erros
        })

    except Exception as e:
        return JsonResponse({'sucesso': False, 'erro': f'Erro ao ler arquivo: {str(e)}'}, status=500)

    
def tela_login(request):
    if request.user.is_authenticated:
        return redirect('homes')
    return render(request, 'login.html')


def normalizar_data(val):
    """ Converte o valor retornado pelo openpyxl em um objeto datetime.date válido """
    if not val:
        return None
    
    # Se o openpyxl já reconheceu como datetime/date
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val

    # Se o Excel mandou a data como número serial (ex: 45130)
    if isinstance(val, (int, float)):
        try:
            return openpyxl.utils.datetime.from_excel(val).date()
        except Exception:
            return None

    # Se veio como string
    if isinstance(val, str):
        val = val.strip().replace('.', '/') # Normaliza 24.07.2026 para 24/07/2026
        
        formatos = [
            "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d",
            "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S"
        ]
        
        for fmt in formatos:
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                pass
                
    return None


def login_usuario(request):
    if request.method == "POST":
        usuario_post = request.POST.get("username")
        senha_post = request.POST.get("password")

        user = authenticate(request, username=usuario_post, password=senha_post)
        if user is not None:
            login(request, user)
            return redirect('homes')
        else:
            messages.error(request, "Usuário ou senha incorretos. Tente novamente.")
            return redirect('tela_login')

    return redirect('tela_login')


def logout_usuario(request):
    logout(request)
    return redirect('tela_login')


def home(request):

    data_atual = timezone.localdate()

    # Atualização automática de contas atrasadas no banco de dados
    ContaPagar.objects.filter(vencimento__lt=data_atual).exclude(status__icontains="Pago").update(status="Atrasado")

    queryset = ContaPagar.objects.all()

    # Captura de Filtros Dinâmicos
    filtro_conciliacao = request.GET.get("conciliacao")
    filtro_data = request.GET.get("data")
    filtro_fornecedor = request.GET.get("fornecedor")
    filtro_categoria = request.GET.get("categoria")
    filtro_banco = request.GET.get("banco")
    filtro_parcela = request.GET.get("parcela")
    filtro_valor = request.GET.get("valor")
    filtro_observacao = request.GET.get("observacao")
    filtro_status = request.GET.get("status")
    filtro_nota_fiscal = request.GET.get("nota_fiscal")
    filtro_linha_digitavel = request.GET.get("linha_digitavel")

    if filtro_conciliacao and filtro_conciliacao.strip():
        queryset = queryset.filter(conciliado__icontains=filtro_conciliacao)
    if filtro_fornecedor and filtro_fornecedor.strip():
        queryset = queryset.filter(fornecedor__icontains=filtro_fornecedor)
    if filtro_categoria and filtro_categoria.strip():
        queryset = queryset.filter(categoria__icontains=filtro_categoria)
    if filtro_banco and filtro_banco.strip():
        queryset = queryset.filter(banco__icontains=filtro_banco)
    if filtro_parcela and filtro_parcela.strip():
        queryset = queryset.filter(parcela__icontains=filtro_parcela)
    if filtro_valor and filtro_valor.strip():
        queryset = queryset.filter(valor__icontains=filtro_valor)
    if filtro_observacao and filtro_observacao.strip():
        queryset = queryset.filter(observacao__icontains=filtro_observacao)
    if filtro_status and filtro_status.strip():
        queryset = queryset.filter(status__icontains=filtro_status)
    if filtro_nota_fiscal and filtro_nota_fiscal.strip():
            queryset = queryset.filter(nota_fiscal__icontains=filtro_nota_fiscal)
    if filtro_linha_digitavel and filtro_linha_digitavel.strip():
            queryset = queryset.filter(linha_digitavel__icontains=filtro_linha_digitavel)

    if filtro_data and filtro_data.strip():
        try:
            data_objeto = datetime.strptime(filtro_data.strip(), "%d/%m/%Y")
            queryset = queryset.filter(vencimento=data_objeto.strftime("%Y-%m-%d"))
        except ValueError:
            messages.error(request, "Formato de data inválido. Use DD/MM/AAAA.")

    paginator = Paginator(queryset, 15)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # Listas usadas pra popular os <select> do modal de edição
    # (fornecedor/categoria/banco no ContaPagar são CharField de texto
    # puro, mas o valor digitado precisa vir padronizado dessas tabelas)
    bancos_disponiveis = BancoSaldo.objects.all().order_by('nome')
    fornecedores = Fornecedor.objects.filter(ativo=True).order_by('razao_social')
    categorias = Categoria.objects.all().order_by('nome')

    context = {
        "page_obj": page_obj,
        "bancos_disponiveis": bancos_disponiveis,  # usado em outros pontos do template
        "fornecedores": fornecedores,
        "categorias": categorias,
        "bancos": bancos_disponiveis,  # mesmo queryset, nome que o modal de edição espera
    }
    return render(request, "home.html", context)


def aba_conciliacao(request):
    # 1. Coletamos a lista de IDs de contas que JÁ FORAM conciliadas
    ids_conciliados = ConciliacaoBancaria.objects.values_list('conta_pagar_id', flat=True).distinct()

    # 2. MODIFICAÇÃO: Removemos o .exclude() para que tudo continue aparecendo na tela!
    queryset = ContaPagar.objects.filter(status__icontains="Pago").order_by("-vencimento")

    # Aplica o filtro de fornecedor se houver
    filtro_fornecedor = request.GET.get("fornecedor")
    if filtro_fornecedor and filtro_fornecedor.strip():
        queryset = queryset.filter(fornecedor__icontains=filtro_fornecedor)

    # 3. Cálculo dos totais direto no banco de dados (Mantido exatamente igual)
    ids_conta_manual = ConciliacaoBancaria.objects.filter(
        transacao_extrato_id=0
    ).values_list('conta_pagar_id', flat=True)
    
    total_manual = ContaPagar.objects.filter(
        id__in=ids_conta_manual
    ).aggregate(
        total=Coalesce(Sum('valor'), Value(0.0), output_field=FloatField())
    )['total']

    ids_extrato_ofx = ConciliacaoBancaria.objects.exclude(
        transacao_extrato_id=0
    ).values_list('transacao_extrato_id', flat=True)
    
    total_ofx = TransacaoExtrato.objects.filter(
        id__in=ids_extrato_ofx
    ).aggregate(
        total=Coalesce(Sum('valor_extrato'), Value(0.0), output_field=FloatField())
    )['total']

    total_conciliado = total_manual + total_ofx

    total_pendente = queryset.exclude(id__in=ids_conciliados).aggregate(
        total=Coalesce(Sum('valor'), Value(0.0), output_field=FloatField())
    )['total']

    # 4. Paginação
    paginator = Paginator(queryset, 15)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    bancos_disponiveis = BancoSaldo.objects.all().order_by('nome')

    context = {
        'page_obj': page_obj,
        'total_conciliado': total_conciliado,
        'total_pendente': total_pendente, 
        'bancos_disponiveis': bancos_disponiveis,
        # Convertemos para set para busca ultra rápida no template com o operador "in"
        'ids_conciliados': set(ids_conciliados), 
        'filtro_fornecedor': filtro_fornecedor,
    }
    return render(request, 'conciliacao.html', context)


def processar_ofx_ajax(request):
    if request.method != "POST" or not request.FILES.get("arquivo_ofx"):
        return JsonResponse({'success': False, 'error': 'Método inválido ou arquivo não enviado.'})

    banco_destino = request.POST.get("banco_destino")
    arquivo_ofx = request.FILES.get("arquivo_ofx")

    if OFXTree is None:
        return JsonResponse({'success': False, 'error': 'Biblioteca ofxtools não está instalada no ambiente.'})

    try:
        # 1. Lê o arquivo vindo do upload diretamente como texto (string)
        conteudo = arquivo_ofx.read().decode("utf-8", errors="ignore")

        # 2. Correção cirúrgica para datas zeradas que quebram o parser
        conteudo_corrigido = conteudo.replace("00000000000000", "20000101120000").replace("00000000", "20000101")

        # 3. Transforma o texto corrigido de volta em bytes na memória e faz o parse
        parser = OFXTree()
        parser.parse(io.BytesIO(conteudo_corrigido.encode("utf-8")))
        obj = parser.convert()

        # 4. Acessa o primeiro extrato de forma direta
        transacoes = obj.statements[0].banktranlist
        transacoes_ofx = []
        datas_transacoes = []  # Lista auxiliar para capturar o período real das transações

        # --- BUSCA DE HISTÓRICO: Pegamos os IDs de transações do extrato que JÁ foram conciliadas ---
        extratos_ja_conciliados = set(
            ConciliacaoBancaria.objects.exclude(
                transacao_extrato_id=0
            ).values_list('transacao_extrato_id', flat=True)
        )

        # Usamos uma transação atômica para salvar tudo de uma vez com máxima performance
        with transaction.atomic():
            # 5. Percorre as transações filtrando estritamente as saídas do extrato (débitos)
            for tx in transacoes:
                # Converte o valor original do extrato para Decimal com segurança
                valor_trnamt = Decimal(str(tx.trnamt))
                if valor_trnamt >= 0:
                    continue  # Entrada (crédito) -> ignora

                valor_ajustado = abs(valor_trnamt)
                
                # --- TRATAMENTO ROBUSTO DE DATA ---
                # Garante que dtposted vire um objeto date puro do Python para evitar TypeErrors
                data_real = None
                if tx.dtposted:
                    if isinstance(tx.dtposted, (datetime, date)):
                        data_real = tx.dtposted if isinstance(tx.dtposted, date) else tx.dtposted.date()
                    else:
                        # Se por acaso o parser trouxe como string, tenta converter
                        try:
                            data_real = datetime.strptime(str(tx.dtposted)[:10], '%Y-%m-%d').date()
                        except ValueError:
                            pass

                if data_real:
                    datas_transacoes.append(data_real)

                # Mantém a formatação segura para gravação e exibição
                data_iso = data_real.strftime('%Y-%m-%d') if data_real else None
                data_formatada = data_real.strftime('%d/%m/%Y') if data_real else '-'

                transacao_banco, _ = TransacaoExtrato.objects.update_or_create(
                    fitid=tx.fitid,
                    defaults={
                        'banco_origem': banco_destino,
                        'data_banco': data_iso,
                        'descricao_ofx': tx.memo if tx.memo else (tx.name or "Transação sem descrição"),
                        'valor_extrato': valor_ajustado,
                    }
                )

                # Verifica se este registro do extrato já possui conciliação
                ja_conciliado = transacao_banco.id in extratos_ja_conciliados

                transacoes_ofx.append({
                    'id': transacao_banco.id,
                    'data': data_formatada,
                    'descricao': transacao_banco.descricao_ofx,
                    'valor': str(valor_ajustado),  # Retorna como string para garantir precisão no JS
                    'ja_conciliado': ja_conciliado,
                })

        # --- FLUXO DE AUDITORIA (D+1) INTELIGENTE COM LIMITE DE DATA ---

        # Query de IDs que já foram conciliados
        ids_ja_conciliados = ConciliacaoBancaria.objects.values_list('conta_pagar_id', flat=True)

        # Filtro base de contas que ainda não foram conciliadas no banco de destino
        contas_filtro = ContaPagar.objects.filter(
            banco=banco_destino,
            status__icontains="Pago",
        ).exclude(
            id__in=ids_ja_conciliados,
        )

        # Se o extrato continha transações com datas válidas, aplicamos a janela de tolerância de data
        if datas_transacoes:
            menor_data_ofx = min(datas_transacoes)
            maior_data_ofx = max(datas_transacoes)

            # Define uma margem de segurança de 7 dias para antes da menor transação e depois da maior.
            data_inicio_limite = menor_data_ofx - timedelta(days=7)
            data_fim_limite = maior_data_ofx + timedelta(days=7)

            # Filtra o vencimento da conta dentro do intervalo correspondente ao período do OFX
            contas_filtro = contas_filtro.filter(
                vencimento__range=(data_inicio_limite, data_fim_limite)
            )

        # Executa a busca otimizada trazendo apenas as colunas essenciais como dicionário
        contas_sistema = contas_filtro.order_by('vencimento').values('id', 'fornecedor', 'valor', 'vencimento')

        # Monta a estrutura de retorno para o frontend
        contas_pendentes = [
            {
                'id': conta['id'],
                'fornecedor': conta['fornecedor'],
                'valor': str(conta['valor']),  # Enviado como string para preservar os centavos
                'data_pagamento': conta['vencimento'].strftime('%d/%m/%Y') if conta['vencimento'] else '-',
            }
            for conta in contas_sistema
        ]

        return JsonResponse({
            'success': True,
            'banco': banco_destino,
            'transacoes_ofx': transacoes_ofx,
            'contas_pendentes': contas_pendentes,
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': f"Erro ao processar OFX: {str(e)}"})


def gravar_conciliacao_lote(request):
    if request.method == "POST":
        try:
            dados = json.loads(request.body)
            vinculos = dados.get("vinculos", [])
            banco_pago = dados.get("banco", "Definir")

            for item in vinculos:
                c_id = int(item['conta_id'])
                e_id = int(item['extrato_id'])

                conta = ContaPagar.objects.get(id=c_id)
                extrato = TransacaoExtrato.objects.get(id=e_id)

                ConciliacaoBancaria.objects.create(
                    conta_pagar_id=conta.id,
                    transacao_extrato_id=extrato.id,
                    data_conciliacao=extrato.data_banco,
                    banco_pago=banco_pago,
                    valor_original_conta=conta.valor,
                    valor_pago_extrato=extrato.valor_extrato,
                )

            return JsonResponse({"success": True})
        except Exception as e:
            return JsonResponse({"success": False, "error": f"Erro na gravação: {str(e)}"})

    return JsonResponse({"success": False, "error": "Método inválido."})


def form(request):
    if request.method == "POST":
        fornecedor = request.POST.get("fornecedor")
        banco_nome = request.POST.get("banco")
        categoria_nome = request.POST.get("categoria")
        parcelas = request.POST.get("parcela")
        valor_str = request.POST.get("valor")
        vencimento_data = request.POST.get("vencimento_manual")

        vencimento_base = datetime.strptime(vencimento_data, "%Y-%m-%d").date()

        if int(parcelas) == 1:
            ContaPagar.objects.create(
                fornecedor=fornecedor,
                banco=banco_nome,
                categoria=categoria_nome,
                parcela=f'0{parcelas}/0{parcelas}',
                valor=parse_valor(valor_str),
                vencimento=vencimento_base,
                status="Pendente",
            )
            return redirect("homes")

        valor_parcela = parse_valor(valor_str)

        for i in range(1, int(parcelas) + 1):
            vencimento_parcela = vencimento_base + relativedelta(months=i - 1)
            ContaPagar.objects.create(
                fornecedor=fornecedor,
                banco=banco_nome,
                categoria=categoria_nome,
                parcela=f'0{i}/0{parcelas}',
                valor=valor_parcela,
                vencimento=vencimento_parcela,
                status="Pendente",
            )
        return redirect("homes")

    fornecedores_reais = Fornecedor.objects.values_list('razao_social', flat=True).distinct().order_by('razao_social')
    bancos_reais = BancoSaldo.objects.values_list('nome', flat=True).distinct().order_by('nome')
    categorias_reais = Categoria.objects.all().order_by('grupo')

    opcoes_parcelas = [f'{i}' for i in range(1, 25)]

    contexto = {
        "fornecedores": list(fornecedores_reais),
        "bancos": list(bancos_reais),
        "categorias": categorias_reais,
        "opcoes_parcelas": opcoes_parcelas,
    }
    return render(request, "form.html", contexto)


def conciliar(request, identi):
    # Verifica se a conta já possui uma conciliação ativa para inverter o status
    if ConciliacaoBancaria.objects.filter(conta_pagar_id=identi).exists():
        ConciliacaoBancaria.objects.filter(conta_pagar_id=identi).delete()
    else:
        conta = get_object_or_404(ContaPagar, id=identi)

        ConciliacaoBancaria.objects.create(
            conta_pagar_id=conta.id,
            transacao_extrato_id=0,  # 0 indica ajuste manual sem arquivo OFX
            data_conciliacao=date.today(),
            banco_pago=conta.banco,
        )

    url_anterior = request.META.get("HTTP_REFERER")
    return redirect(url_anterior) if url_anterior else redirect("homes")


def atualizar_status_json(request, identi):
    if request.method == 'POST':
        conta = get_object_or_404(ContaPagar, id=identi)
        novo_status = request.POST.get('status')
        nova_data = request.POST.get('ultimo_pagamento')
        novo_juros = request.POST.get('juros')
        nova_conta_origem = request.POST.get('conta_origem')

        if novo_status:
            conta.status = novo_status
        conta.ultimo_pagamento = nova_data if nova_data else None
        if novo_juros:
            conta.juros = novo_juros
        if nova_conta_origem:
            conta.banco_pago = nova_conta_origem

        conta.save()
        return JsonResponse({'success': True})

    return JsonResponse({'success': False}, status=400)


def provisao_periodo(request):
    hoje = timezone.localdate()
    futuro_padrao = hoje + timedelta(days=30)

    data_inicio_str = request.GET.get("data_inicio")
    data_fim_str = request.GET.get("data_fim")

    data_inicio = hoje
    data_fim = futuro_padrao

    if data_inicio_str:
        try:
            data_inicio = datetime.strptime(data_inicio_str, "%Y-%m-%d").date()
        except ValueError:
            pass
    if data_fim_str:
        try:
            data_fim = datetime.strptime(data_fim_str, "%Y-%m-%d").date()
        except ValueError:
            pass

    contas_periodo = ContaPagar.objects.filter(
        vencimento__range=(data_inicio, data_fim)
    ).exclude(status__icontains="Pago").order_by("vencimento")

    soma_total = contas_periodo.aggregate(Sum("valor"))["valor__sum"] or 0.00
    total_registros = contas_periodo.count()
    bancos_disponiveis = BancoSaldo.objects.all().order_by('nome')

    # Listas usadas pra popular os <select> do modal de editar registro
    fornecedores = Fornecedor.objects.filter(ativo=True).order_by('razao_social')
    categorias = Categoria.objects.all().order_by('nome')

    context = {
        "contas": contas_periodo,
        "total_valor": soma_total,
        "total_registros": total_registros,
        "data_inicio": data_inicio.strftime("%Y-%m-%d"),
        "data_fim": data_fim.strftime("%Y-%m-%d"),
        "bancos_disponiveis": bancos_disponiveis,
        "fornecedores": fornecedores,
        "categorias": categorias,
        "bancos": bancos_disponiveis,
    }
    return render(request, "provisao.html", context)


def cadastrar_fornecedor(request):
    if request.method == 'POST':
        razao_social = request.POST.get('razao_social')
        nome_fantasia = request.POST.get('nome_fantasia')
        cnpj = request.POST.get('cnpj')
        email = request.POST.get('email')
        telefone = request.POST.get('telefone')
        logradouro = request.POST.get('logradouro')
        cidade = request.POST.get('cidade')
        estado = request.POST.get('estado')

        if not razao_social or not cnpj:
            messages.error(request, "Razão Social e CNPJ são obrigatórios.")
            return render(request, 'cadastrar_fornecedor.html', {'dados': request.POST})

        if Fornecedor.objects.filter(cnpj=cnpj).exists():
            messages.error(request, "Este CNPJ já está cadastrado.")
            return render(request, 'cadastrar_fornecedor.html', {'dados': request.POST})

        try:
            Fornecedor.objects.create(
                razao_social=razao_social, nome_fantasia=nome_fantasia, cnpj=cnpj,
                email=email, telefone=telefone, logradouro=logradouro, cidade=cidade, estado=estado,
            )
            messages.success(request, f"Fornecedor '{nome_fantasia or razao_social}' cadastrado com sucesso!")
            return redirect('homes')
        except Exception as e:
            messages.error(request, f"Erro ao cadastrar fornecedor: {e}")

    return render(request, 'cadastrar_fornecedor.html')


def saldo(request):

    bancos = BancoSaldo.objects.all().order_by('nome')

    dados = FechamentoCaixa.objects.all().values_list(
    'data',        
    'unidade', 
    'abertura', 
    'suprimento', 
    'saidas', 
    'troco', 
    'vendas', 
    'total_dinheiro', 
    'sangria')

    df = pd.DataFrame(
        list(dados), 
        columns=['Data', 'Unidade', 'Abertura', 'Suprimento', 'Saídas', 'Troco', 'Vendas', 'Total Dinheiro', 'Sangria']
    )

    df['Unidade'] = df['Unidade'].str.strip()

    total = df['Sangria'].sum()

    saldo_aero = df[df['Unidade'] == 'AEROPORTO']['Sangria'].sum()

    saldo_casta = df[df['Unidade'] == 'BR 316']['Sangria'].sum()

    saldo_baenoso = df[df['Unidade'] == 'ESTADIO BAENAO']['Sangria'].sum()

    saldo_patio = df[df['Unidade'] == 'PADRE EUTIQUIO']['Sangria'].sum()

    context = {'total':total, 'saldo_aero': saldo_aero, 'saldo_casta': saldo_casta, 'saldo_baenoso': saldo_baenoso, 'saldo_patio': saldo_patio, 'bancos': bancos}

    return render(request, "saldo.html", context)


def dashboard_leve(request):
    hoje = date.today()
    ids_conciliados = ConciliacaoBancaria.objects.values_list('conta_pagar_id', flat=True).distinct()

    metricas = ContaPagar.objects.aggregate(
        total_pagos=Count('id', filter=Q(status="Pago")),
        volume_atrasado=Sum('valor', filter=Q(status="Pendente", vencimento__lt=hoje)),
    )

    total_juros = sum(c.juros for c in ConciliacaoBancaria.objects.all() if hasattr(c, 'juros'))
    volume_atrasado = metricas['volume_atrasado'] or 0
    total_pagos = metricas['total_pagos'] or 1

    pagos_conciliados_count = len(ids_conciliados)
    taxa_conciliacao = (pagos_conciliados_count / total_pagos) * 100

    dados_grafico = ConciliacaoBancaria.objects.all()
    categoria_dict = {}
    for c in dados_grafico:
        juros_val = float(getattr(c, 'juros', 0) or 0)
        if juros_val > 0:
            conta = ContaPagar.objects.filter(id=c.conta_pagar_id).first()
            cat_nome = conta.categoria if conta else "Sem Categoria"
            categoria_dict[cat_nome] = categoria_dict.get(cat_nome, 0) + juros_val

    categorias = list(categoria_dict.keys())
    juros_valores = list(categoria_dict.values())

    pendentes_conciliacao = ContaPagar.objects.filter(status="Pago").exclude(id__in=ids_conciliados)[:5]

    context = {
        'total_juros': total_juros,
        'taxa_conciliacao': round(taxa_conciliacao, 1),
        'volume_atrasado': volume_atrasado,
        'categorias_json': json.dumps(categorias),
        'juros_json': json.dumps(juros_valores),
        'pendentes_conciliacao': pendentes_conciliacao,
    }
    return render(request, 'dashboard.html', context)


def salvar_conciliacao_lote(request):
    if request.method == "POST":
        try:
            dados = json.loads(request.body)
            vinculos = dados.get("vinculos", [])

            if not vinculos:
                return JsonResponse({'success': False, 'error': 'Nenhum vínculo selecionado.'})

            agora = timezone.now()

            for item in vinculos:
                conta_id = item.get("conta_id")
                extrato_id = item.get("extrato_id")

                ConciliacaoBancaria.objects.get_or_create(
                    conta_pagar_id=conta_id,
                    defaults={
                        'transacao_extrato_id': extrato_id,
                        'data_conciliacao': agora,
                    }
                )
                # Se no seu fluxo precisar marcar a conta como liquidada/auditada:
                # ContaPagar.objects.filter(id=conta_id).update(status="Pago e Conciliado")

            return JsonResponse({'success': True})

        except Exception as e:
            return JsonResponse({'success': False, 'error': f"Erro interno ao salvar lote: {str(e)}"})

    return JsonResponse({'success': False, 'error': 'Método não permitido.'})


def baixar_planilha_padrao(request):
    workbook = Workbook()
    aba = workbook.active
    aba.title = "Modelo"

    cabecalho = ["vencimento", "fornecedor", "categoria", "banco", "parcela", "valor", "observação", "status", "nota_fiscal", "linha_digitavel"]
    aba.append(cabecalho)
    aba.append(["12/12/2026", "Seu Fornecedor", "Energia", "Nome Banco", "1/1", 150.00, "Sua Observação", "Pendente","0000", "00000000000000000000000000000"])

    for coluna in aba.columns:
        maior_largura = max(len(str(celula.value)) for celula in coluna)
        aba.column_dimensions[coluna[0].column_letter].width = maior_largura + 4

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="modelo_contas_a_pagar.xlsx"'
    workbook.save(response)
    return response


def atualizar_registro(request):
    # Volta pra página de onde o form foi enviado (home, provisao, etc.),
    # em vez de sempre mandar pra "homes" -- mesmo padrão já usado em
    # conciliar().
    url_anterior = request.META.get("HTTP_REFERER")
    destino = url_anterior if url_anterior else redirect('homes').url

    if request.method != "POST":
        return redirect(destino)

    registro_id = request.POST.get('id')
    registro = get_object_or_404(ContaPagar, id=registro_id)

    campos_alterados = []

    vencimento_str = request.POST.get('vencimento', '').strip()
    if vencimento_str:
        try:
            nova_vencimento = datetime.strptime(vencimento_str, "%Y-%m-%d").date()
        except ValueError:
            messages.error(request, "Data de vencimento inválida.")
            return redirect(destino)
        if nova_vencimento != registro.vencimento:
            registro.vencimento = nova_vencimento
            campos_alterados.append('vencimento')

    novo_fornecedor = request.POST.get('fornecedor', '').strip()
    if novo_fornecedor and novo_fornecedor != registro.fornecedor:
        registro.fornecedor = novo_fornecedor
        campos_alterados.append('fornecedor')

    nova_categoria = request.POST.get('categoria', '').strip()
    if nova_categoria != (registro.categoria or ''):
        registro.categoria = nova_categoria or None
        campos_alterados.append('categoria')

    novo_banco = request.POST.get('banco', '').strip()
    if novo_banco and novo_banco != registro.banco:
        registro.banco = novo_banco
        campos_alterados.append('banco')

    nova_parcela = request.POST.get('parcela', '').strip()
    if nova_parcela and nova_parcela != registro.parcela:
        registro.parcela = nova_parcela
        campos_alterados.append('parcela')

    nova_observacao = request.POST.get('observacao', '').strip()
    if nova_observacao != (registro.observacao or ''):
        registro.observacao = nova_observacao
        campos_alterados.append('observacao')

    valor_str = request.POST.get('valor', '').strip()
    if valor_str:
        try:
            novo_valor = parse_valor(valor_str)
        except InvalidOperation:
            messages.error(request, f"Valor inválido: “{valor_str}”.")
            return redirect(destino)
        if novo_valor != registro.valor:
            registro.valor = novo_valor
            campos_alterados.append('valor')

    if campos_alterados:
        registro.save(update_fields=campos_alterados)
        messages.success(request, "Registro atualizado com sucesso.")
    else:
        messages.info(request, "Nenhuma alteração foi detectada.")

    return redirect(destino)


def importar_fechamento_caixa(request):

    print("\n" + "="*50)
    print("1. INICIANDO IMPORTAÇÃO COM OPENPYXL")
    print("="*50)

    if request.method == 'POST':
        if 'arquivo' not in request.FILES:
            print("❌ ERRO: Nenhum arquivo enviado na requisição.")
            messages.error(request, 'Nenhum arquivo foi selecionado.')
            return redirect(request.META.get('HTTP_REFERER', '/'))

        arquivo = request.FILES['arquivo']
        print(f"2. Arquivo recebido: {arquivo.name}")

        # Valida extensão
        if not arquivo.name.endswith(('.xlsx', '.xlsm')):
            print("❌ ERRO: Formato inválido. openpyxl lê apenas .xlsx / .xlsm")
            messages.error(request, 'Envie um arquivo do Excel (.xlsx). Para CSV, salve em .xlsx.')
            return redirect(request.META.get('HTTP_REFERER', '/'))

        try:
            # Carrega a planilha na memória
            wb = openpyxl.load_workbook(arquivo, data_only=True)
            sheet = wb.active
            print(f"3. Planilha aberta com sucesso! Aba ativa: {sheet.title}")

            # Pega o cabeçalho (primeira linha) e normaliza
            primeira_linha = [str(cell.value or '').strip().lower() for cell in sheet[1]]
            print("Cabeçalhos encontrados:", primeira_linha)

            # Mapeia qual coluna é qual pelo nome do cabeçalho
            def get_col_index(nome_coluna):
                try:
                    return primeira_linha.index(nome_coluna)
                except ValueError:
                    return None

            col_data = get_col_index('data')
            col_unidade = get_col_index('unidade')
            col_abertura = get_col_index('abertura')
            col_suprimento = get_col_index('suprimento')
            col_saidas = get_col_index('saidas') or get_col_index('saídas')
            col_troco = get_col_index('troco')
            col_vendas = get_col_index('vendas')
            col_total_dinheiro = get_col_index('total dinheiro')
            col_sangria = get_col_index('sangria')

            if col_data is None:
                print("❌ ERRO: Coluna 'data' não foi encontrada na planilha!")
                messages.error(request, "Cabeçalho 'Data' não encontrado na planilha.")
                return redirect(request.META.get('HTTP_REFERER', '/'))

            registros = []

            # Percorre as linhas a partir da segunda (pula o cabeçalho)
            for row in sheet.iter_rows(min_row=2, values_only=True):
                # Pula linha totalmente vazia
                if not any(row):
                    continue

                # Trata a Data
                val_data = row[col_data] if col_data is not None else None
                data_formatada = None

                if isinstance(val_data, datetime):
                    data_formatada = val_data.date()
                elif isinstance(val_data, str) and val_data.strip():
                    try:
                        # Tenta converter texto em data (formato DD/MM/AAAA)
                        data_formatada = datetime.strptime(val_data.strip(), "%d/%m/%Y").date()
                    except ValueError:
                        try:
                            # Tenta converter texto em data (formato AAAA-MM-DD)
                            data_formatada = datetime.strptime(val_data.strip(), "%Y-%m-%d").date()
                        except ValueError:
                            pass

                if not data_formatada:
                    print(f"⚠️ Pulei linha por data inválida: {val_data}")
                    continue

                # Função auxiliar para limpar floats/decimals de forma segura
                def clean_decimal(val):
                    if val is None:
                        return 0.0
                    try:
                        return float(val)
                    except (ValueError, TypeError):
                        return 0.0

                registros.append(
                    FechamentoCaixa(
                        data=data_formatada,
                        unidade=str(row[col_unidade]) if col_unidade is not None and row[col_unidade] else '',
                        abertura=clean_decimal(row[col_abertura] if col_abertura is not None else 0),
                        suprimento=clean_decimal(row[col_suprimento] if col_suprimento is not None else 0),
                        saidas=clean_decimal(row[col_saidas] if col_saidas is not None else 0),
                        troco=clean_decimal(row[col_troco] if col_troco is not None else 0),
                        vendas=clean_decimal(row[col_vendas] if col_vendas is not None else 0),
                        total_dinheiro=clean_decimal(row[col_total_dinheiro] if col_total_dinheiro is not None else 0),
                        sangria=clean_decimal(row[col_sangria] if col_sangria is not None else 0),
                    )
                )

            print(f"4. Total de linhas válidas processadas: {len(registros)}")

            if registros:
                FechamentoCaixa.objects.bulk_create(registros)
                print("5. 🚀 SUCESSO! Dados salvos no banco!")
                messages.success(request, f'{len(registros)} registros importados com sucesso!')
            else:
                print("⚠️ AVISO: Nenhuma linha com data válida foi convertida.")
                messages.warning(request, 'Nenhum dado válido encontrado na planilha.')

        except Exception as e:
            print(f"❌ ERRO EXCEÇÃO: {str(e)}")
            messages.error(request, f'Erro ao ler o arquivo Excel: {str(e)}')

        return redirect(request.META.get('HTTP_REFERER', '/'))

    return redirect('/')


def deposito(request):
    if request.method == 'POST':
        # Captura os dados enviados pelo formulário manual
        unidade = request.POST.get('unidade')
        data_deposito = request.POST.get('data_deposito')
        valor = request.POST.get('valor')
        destino = request.POST.get('destino')
        comprovante = request.FILES.get('comprovante')
        observacao = request.POST.get('observacao')

        # Validação simples de campos obrigatórios
        if unidade and data_deposito and valor:
            Deposito.objects.create(
                unidade=unidade,
                data_deposito=data_deposito,
                valor=valor,
                destino=destino,
                comprovante=comprovante,
                observacao=observacao
            )
            messages.success(request, 'Depósito registrado com sucesso!')
        else:
            messages.error(request, 'Erro ao registrar: preencha todos os campos obrigatórios.')

    # Redireciona de volta para a página anterior/saldo
    return redirect(request.META.get('HTTP_REFERER', '/'))


def exportar_contas_pagar_excel(request):
    tipo_exportacao = request.GET.get('tipo', 'tudo')
    data_inicio = request.GET.get('data_inicio')
    data_fim = request.GET.get('data_fim')

    # Queryset ordenado por vencimento
    queryset = ContaPagar.objects.all().order_by('vencimento')

    if tipo_exportacao == 'periodo' and data_inicio and data_fim:
        queryset = queryset.filter(vencimento__range=[data_inicio, data_fim])

    # Instância do Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contas a Pagar"
    ws.views.sheetView[0].showGridLines = True

    # --- ESTILOS VISUAIS ---
    font_header = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    fill_header = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')

    font_dados = Font(name='Segoe UI', size=10)
    fill_zebra = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
    fill_white = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

    font_total = Font(name='Segoe UI', size=11, bold=True, color='1F497D')
    fill_total = PatternFill(start_color='DCE6F1', end_color='DCE6F1', fill_type='solid')

    border_thin = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    border_total = Border(
        top=Side(style='thin', color='1F497D'),
        bottom=Side(style='double', color='1F497D')
    )

    align_center = Alignment(horizontal='center', vertical='center')
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    # --- LINHA 1: CABEÇALHO COMPLETO ---
    headers = [
        'ID', 'Status', 'Vencimento', 'Fornecedor', 'Categoria', 
        'Nota Fiscal', 'Parcela', 'Valor (R$)', 'Banco', 'Linha Digitável', 'Observação'
    ]
    ws.append(headers)

    # Estiliza o cabeçalho (Linha 1)
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_thin

    # --- LINHAS DE DADOS (Iniciando na Linha 2) ---
    start_row = 2
    for idx, conta in enumerate(queryset, start=start_row):
        is_even = (idx - start_row) % 2 == 0
        current_fill = fill_white if is_even else fill_zebra

        venc = getattr(conta, 'vencimento', None)
        venc_str = venc.strftime('%d/%m/%Y') if venc else '-'

        valor = float(getattr(conta, 'valor', 0) or 0)

        # Preenchimento de todas as colunas do model
        row_data = [
            conta.id,
            str(getattr(conta, 'status', '-')).upper(),
            venc_str,
            str(getattr(conta, 'fornecedor', '-')),
            str(getattr(conta, 'categoria', '-')),
            str(getattr(conta, 'nota_fiscal', '-')),
            str(getattr(conta, 'parcela', '-')),
            valor,
            str(getattr(conta, 'banco', '-')),
            str(getattr(conta, 'linha_digitavel', '-')),
            str(getattr(conta, 'observacao', '-'))
        ]

        ws.append(row_data)

        # Formatação célula a célula
        for col_num in range(1, len(headers) + 1):
            c = ws.cell(row=idx, column=col_num)
            c.font = font_dados
            c.fill = current_fill
            c.border = border_thin

            # Alinhamentos específicos
            if col_num in (1, 2, 3, 6, 7):  # ID, Status, Vencimento, Nota Fiscal, Parcela
                c.alignment = align_center
            elif col_num in (4, 5, 9, 10, 11):  # Textos / Observações
                c.alignment = align_left
            elif col_num == 8:  # Valor
                c.alignment = align_right
                c.number_format = 'R$ #,##0.00'

    last_row = start_row + len(queryset) - 1 if len(queryset) > 0 else start_row

    # --- LINHA DE TOTAIS (No final dos dados) ---
    if len(queryset) > 0:
        total_row = last_row + 1
        
        ws.cell(row=total_row, column=1, value="TOTAL")
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=7)
        ws.cell(row=total_row, column=1).alignment = Alignment(horizontal='right', vertical='center')

        # Soma automática na coluna de Valor (Coluna H / 8)
        sum_cell = ws.cell(row=total_row, column=8, value=f"=SUM(H{start_row}:H{last_row})")
        sum_cell.number_format = 'R$ #,##0.00'
        sum_cell.alignment = align_right

        for col_num in range(1, len(headers) + 1):
            c = ws.cell(row=total_row, column=col_num)
            c.font = font_total
            c.fill = fill_total
            c.border = border_total

    # --- AJUSTES FINAIS ---
    # Largura automática das colunas
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # Congela apenas a linha 1 (cabeçalho sempre visível ao rolar)
    ws.freeze_panes = 'A2'

    # Resposta para download
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"contas_a_pagar_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response